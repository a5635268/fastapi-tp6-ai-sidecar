"""
Excel 导入导出工具类
提供通用的 Excel 文件处理能力
支持数据导出、导入和模板生成
"""
import io
from datetime import datetime
from typing import Any, Dict, List, Optional, Type, Union, Sequence

from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelUtil:
    """
    Excel 导入导出工具类

    提供数据导出、导入和模板生成功能
    使用 openpyxl 处理 Excel 文件
    """

    # 默认样式配置
    DEFAULT_HEADER_FONT = Font(bold=True, size=12)
    DEFAULT_HEADER_FILL = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    DEFAULT_HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
    DEFAULT_CELL_ALIGNMENT = Alignment(horizontal='left', vertical='center')
    DEFAULT_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    @classmethod
    def export_excel(
        cls,
        data: Sequence[Any],
        columns: Optional[Dict[str, str]] = None,
        filename: str = 'export.xlsx',
        sheet_name: str = 'Sheet1',
        header_style: bool = True,
    ) -> StreamingResponse:
        """
        将数据列表导出为 Excel 文件

        Args:
            data: 数据列表（Pydantic 模型或字典）
            columns: 列映射字典 {字段名: 表头名}，为 None 时自动推断
            filename: 导出文件名
            sheet_name: 工作表名称
            header_style: 是否应用表头样式

        Returns:
            StreamingResponse: 文件下载响应

        Example:
            >>> data = [{'id': 1, 'name': 'Alice'}, {'id': 2, 'name': 'Bob'}]
            >>> columns = {'id': 'ID', 'name': '姓名'}
            >>> return ExcelUtil.export_excel(data, columns, 'users.xlsx')
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 确定列映射
        if columns is None:
            columns = cls._infer_columns(data)

        # 写入表头
        headers = list(columns.values())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            if header_style:
                cell.font = cls.DEFAULT_HEADER_FONT
                cell.fill = cls.DEFAULT_HEADER_FILL
                cell.alignment = cls.DEFAULT_HEADER_ALIGNMENT
                cell.border = cls.DEFAULT_BORDER

        # 写入数据
        field_names = list(columns.keys())
        for row_idx, item in enumerate(data, start=2):
            for col_idx, field_name in enumerate(field_names, start=1):
                value = cls._get_field_value(item, field_name)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cls.DEFAULT_CELL_ALIGNMENT
                cell.border = cls.DEFAULT_BORDER

        # 自动调整列宽
        cls._auto_adjust_column_width(ws, headers)

        # 生成文件流
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        return StreamingResponse(
            content=file_stream,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'},
        )

    @classmethod
    def import_excel(
        cls,
        file_content: bytes,
        model: Optional[Type[BaseModel]] = None,
        columns: Optional[Dict[str, str]] = None,
        skip_header: bool = True,
        start_row: int = 1,
        start_col: int = 1,
    ) -> List[Dict[str, Any]]:
        """
        从 Excel 文件导入数据

        Args:
            file_content: Excel 文件内容（字节）
            model: Pydantic 模型类，用于数据校验（可选）
            columns: 列映射字典 {表头名: 字段名}，为 None 时使用表头名作为字段名
            skip_header: 是否跳过表头行
            start_row: 数据起始行（1-based）
            start_col: 数据起始列（1-based）

        Returns:
            List[Dict]: 导入的数据列表

        Example:
            >>> with open('import.xlsx', 'rb') as f:
            >>>     data = ExcelUtil.import_excel(f.read(), columns={'姓名': 'name', '年龄': 'age'})
        """
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active

        # 确定列映射
        if columns is None and skip_header:
            # 从表头推断列映射
            columns = cls._infer_columns_from_header(ws, start_row, start_col)

        data = []
        row_start = start_row + 1 if skip_header else start_row

        for row_idx in range(row_start, ws.max_row + 1):
            row_data = {}

            if columns:
                for col_idx, (header_name, field_name) in enumerate(columns.items(), start=start_col):
                    value = ws.cell(row=row_idx, column=col_idx).value
                    row_data[field_name] = cls._convert_value(value)
            else:
                for col_idx in range(start_col, ws.max_column + 1):
                    header_name = ws.cell(row=start_row, column=col_idx).value
                    if header_name:
                        value = ws.cell(row=row_idx, column=col_idx).value
                        row_data[str(header_name)] = cls._convert_value(value)

            # 如果提供了 Pydantic 模型，进行数据校验
            if model:
                try:
                    validated = model(**row_data)
                    data.append(validated.model_dump())
                except Exception:
                    # 校验失败时保留原始数据
                    data.append(row_data)
            else:
                data.append(row_data)

        return data

    @classmethod
    def get_excel_template(
        cls,
        columns: Dict[str, str],
        filename: str = 'template.xlsx',
        sheet_name: str = 'Sheet1',
        example_data: Optional[List[Dict[str, Any]]] = None,
    ) -> StreamingResponse:
        """
        生成 Excel 模板文件

        Args:
            columns: 列映射字典 {字段名: 表头名}
            filename: 模板文件名
            sheet_name: 工作表名称
            example_data: 示例数据（可选）

        Returns:
            StreamingResponse: 模板文件下载响应

        Example:
            >>> columns = {'name': '姓名', 'age': '年龄'}
            >>> return ExcelUtil.get_excel_template(columns, 'user_template.xlsx')
        """
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # 写入表头
        headers = list(columns.values())
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = cls.DEFAULT_HEADER_FONT
            cell.fill = cls.DEFAULT_HEADER_FILL
            cell.alignment = cls.DEFAULT_HEADER_ALIGNMENT
            cell.border = cls.DEFAULT_BORDER

        # 写入示例数据（如果有）
        if example_data:
            field_names = list(columns.keys())
            for row_idx, item in enumerate(example_data, start=2):
                for col_idx, field_name in enumerate(field_names, start=1):
                    value = item.get(field_name, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = cls.DEFAULT_CELL_ALIGNMENT
                    cell.border = cls.DEFAULT_BORDER

        # 自动调整列宽
        cls._auto_adjust_column_width(ws, headers)

        # 生成文件流
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)

        return StreamingResponse(
            content=file_stream,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'},
        )

    @classmethod
    def _infer_columns(cls, data: Sequence[Any]) -> Dict[str, str]:
        """
        从数据推断列映射

        Args:
            data: 数据列表

        Returns:
            Dict[str, str]: 列映射字典
        """
        if not data:
            return {}

        first_item = data[0]

        if isinstance(first_item, BaseModel):
            return {field: field for field in first_item.model_fields.keys()}

        if isinstance(first_item, dict):
            return {key: key for key in first_item.keys()}

        return {}

    @classmethod
    def _infer_columns_from_header(cls, ws: Any, start_row: int, start_col: int) -> Dict[str, str]:
        """
        从 Excel 表头推断列映射

        Args:
            ws: 工作表对象
            start_row: 表头行
            start_col: 起始列

        Returns:
            Dict[str, str]: 列映射字典 {表头名: 表头名}
        """
        columns = {}
        for col_idx in range(start_col, ws.max_column + 1):
            header = ws.cell(row=start_row, column=col_idx).value
            if header:
                columns[str(header)] = str(header)
        return columns

    @classmethod
    def _get_field_value(cls, item: Any, field_name: str) -> Any:
        """
        获取字段值

        Args:
            item: 数据项（模型或字典）
            field_name: 字段名

        Returns:
            Any: 字段值
        """
        if isinstance(item, BaseModel):
            return getattr(item, field_name, None)

        if isinstance(item, dict):
            return item.get(field_name)

        return None

    @classmethod
    def _convert_value(cls, value: Any) -> Any:
        """
        转换 Excel 值为 Python 类型

        Args:
            value: Excel 原始值

        Returns:
            Any: 转换后的值
        """
        if value is None:
            return None

        if isinstance(value, datetime):
            return value

        return value

    @classmethod
    def _auto_adjust_column_width(cls, ws: Any, headers: List[str]) -> None:
        """
        自动调整列宽

        Args:
            ws: 工作表对象
            headers: 表头列表
        """
        for col_idx, header in enumerate(headers, start=1):
            # 计算列宽：表头长度 * 2 + 额外空间
            width = max(len(str(header)) * 2, 10)
            ws.column_dimensions[get_column_letter(col_idx)].width = width